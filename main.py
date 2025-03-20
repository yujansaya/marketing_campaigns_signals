import concurrent.futures
import feedparser
import email.utils
from datetime import datetime, timedelta, timezone
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import logging
from urllib.parse import quote_plus

from company_names_graph import run_graph
from llm import choose_relevant_niches
import streamlit as st
import pandas as pd
import io
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from niche_enrichment import *
from snowflake_df_cleaner import clean_company_list
# from st_aggrid import AgGrid

NITTER_INSTANCE = "https://nitter.net"
TRUMP_LINK = "https://truthsocial.com/@realDonaldTrump"

# ✅ Move set_page_config to be the first Streamlit command
st.set_page_config(page_title="MarketMuse – Your AI-powered muse for market inspiration",
                   page_icon="📰",
                   # layout="wide"
                   )


def get_chrome_options(headless=True):
    """Create and return ChromeOptions with pre-configured settings."""
    options = webdriver.ChromeOptions()

    # # Set DNS over HTTPS configuration
    # local_state = {
    #     "dns_over_https.mode": "secure",
    #     "dns_over_https.templates": "https://chrome.cloudflare-dns.com/dns-query",
    # }
    # options.add_experimental_option('localState', local_state)

    # Configure user agent
    user_agent = ('Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/60.0.3112.50 Safari/537.36')
    options.add_argument(f'user-agent={user_agent}')

    if headless:
        options.add_argument("--headless")

    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")  # Required for Streamlit Cloud
    options.add_argument("--disable-dev-shm-usage")  # Use disk instead of memory
    options.add_argument("--disable-extensions")  # Reduce memory usage
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--window-size=1280,800")  # Avoid viewport errors
    return options


def create_driver(options):
    """Initialize and return a new Chrome webdriver."""
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def scrape_nitter(keyword, max_tweets=10):
    """
    Scrape Nitter search results using Selenium.

    Args:
        keyword (str): Search keyword.
        max_tweets (int): Maximum number of tweets to retrieve.

    Returns:
        list: List of tweet texts.
    """
    encoded_keyword = quote_plus(keyword)
    search_url = f"{NITTER_INSTANCE}/search?f=tweets&q={encoded_keyword}+usa"

    options = get_chrome_options(headless=True)
    driver = create_driver(options)

    tweets = []
    st.write(f"Scraping tweets for: **{keyword}**")

    try:
        driver.get(search_url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.tweet-content"))
        )
        tweet_elements = driver.find_elements(By.CSS_SELECTOR, "div.tweet-content")

        tweet_display = st.empty()  # Placeholder for updating UI
        for i, tweet in enumerate(tweet_elements):
            if i >= max_tweets:
                break
            tweets.append(tweet.text)
            tweet_display.write(f"✅ {i + 1}/{max_tweets} tweets fetched...")  # Live progress update
            time.sleep(0.5)  # Simulate live updates

    except Exception as e:
        st.error(f"Error scraping Nitter: {e}")
        logging.error(f"Error scraping Nitter: {e}")
    finally:
        driver.quit()

    return tweets


def fetch_feed(query, days=5):
    """
    Fetch the RSS feed for a given query and filter out news older than 'days' days.

    Args:
        query (str): The search query.
        days (int): Number of days to look back.

    Returns:
        tuple: (query, list of filtered feed entries)
    """
    # Use quote_plus to safely encode the query
    encoded_query = quote_plus(query)
    rss_url = f"https://news.google.com/rss/search?q={encoded_query}+usa"
    feed = feedparser.parse(rss_url)
    filtered_entries = []

    current_time = datetime.now(timezone.utc)
    cutoff_time = current_time - timedelta(days=days)

    st.write(f"Fetching news for: **{query}**")
    news_display = st.empty()  # Placeholder for updating UI

    for entry in feed.entries:
        try:
            published_date = email.utils.parsedate_to_datetime(entry.published)
        except Exception as e:
            st.warning(f"Could not parse date for entry: {entry.get('title', 'No Title')}")
            continue

        if published_date > cutoff_time:
            filtered_entries.append({
                "title": entry.title,
                "link": entry.link,
                "date": entry.published
            })
            news_display.write(f"📰 {len(filtered_entries)} articles fetched...")  # Live update
            time.sleep(0.5)  # Simulate processing delay

    return query, filtered_entries


def scroll_up_until_elements(driver, selector, min_count=10, max_scrolls=15):
    """
    Scrolls the page until at least min_count unique elements (by aria-label) are found.

    Args:
        driver (webdriver): Selenium webdriver instance.
        selector (str): CSS selector to find elements.
        min_count (int): Minimum number of unique elements required.
        max_scrolls (int): Maximum number of scroll attempts.

    Returns:
        list: List of unique aria-label texts.
    """
    unique_texts = set()
    body = driver.find_element(By.TAG_NAME, "body")

    for _ in range(max_scrolls):
        elements = driver.find_elements(By.CSS_SELECTOR, selector)
        for element in elements:
            aria_label = element.get_attribute("aria-label")
            if aria_label:
                unique_texts.add(aria_label)
        if len(unique_texts) >= min_count:
            break
        body.send_keys(Keys.PAGE_DOWN)
        time.sleep(2)  # Allow time for new elements to load
    return list(unique_texts)


def trump_scraper():
    """
    Scrapes posts from Donald Trump's Truth Social page using Selenium.

    Returns:
        list: List of post aria-label texts.
    """
    options = get_chrome_options(headless=True)
    driver = create_driver(options)
    posts = []
    try:
        driver.get(TRUMP_LINK)
        # Wait until the timeline element is loaded
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#timeline"))
        )
        posts = scroll_up_until_elements(driver, "#timeline .status[aria-label]", min_count=10)
    except Exception as e:
        logging.error(f"Error scraping Trump page: {e}")
    finally:
        driver.quit()
    return posts


def convert_json_to_csv(json_data):
    data_list = []
    logging.info(f"Json File: {json_data}")
    # Process "List of Affected Business Categories"
    if "List of Affected Business Categories" in json_data:
        for item in json_data["List of Affected Business Categories"]:
            data_list.append({
                "Business Category Name": item["Business Category Name"],
                "NAIC Code": item["NAIC Code"],
                "Affected Commodities": ", ".join(item["Affected Commodities"]),
                "Potential Impact": item["Potential Impact"]
            })

    # Convert to DataFrame
    df = pd.DataFrame(data_list)

    # Convert DataFrame to CSV format (string buffer)
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)

    # Append two new lines and the summary
    summary_text = json_data.get("Summary of Key Findings", "")
    csv_buffer.write("\n\nSummary of Key Findings:\n")
    csv_buffer.write(f'"{summary_text}"\n')

    # Encode the final CSV content
    csv_data = csv_buffer.getvalue().encode("utf-8")

    return csv_data


def save_scrapes_to_excel(combined_data):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Process news_feeds: Flatten data from each category
        news_rows = []
        for category, items in combined_data.get("news_feeds", {}).items():
            for item in items:
                news_rows.append({
                    "category": category,
                    "title": item.get("title"),
                    "link": item.get("link"),
                    "date": item.get("date")
                })
        df_news = pd.DataFrame(news_rows)
        df_news.to_excel(writer, sheet_name="News Feeds", index=False)

        # Process x_tweets: Each category contains tweet strings
        tweets_rows = []
        for category, tweets in combined_data.get("x_tweets", {}).items():
            for tweet in tweets:
                tweets_rows.append({
                    "category": category,
                    "tweet": tweet
                })
        df_tweets = pd.DataFrame(tweets_rows)
        df_tweets.to_excel(writer, sheet_name="Tweets", index=False)

        # Process trump_data: Each category contains tweet strings
        trump_rows = []
        for category, tweets in combined_data.get("trump_data", {}).items():
            for tweet in tweets:
                trump_rows.append({
                    "category": category,
                    "tweet": tweet
                })
        df_trump = pd.DataFrame(trump_rows)
        df_trump.to_excel(writer, sheet_name="Trump Tweets", index=False)

    # Load the workbook to modify it after writing
    output.seek(0)
    workbook = load_workbook(output)

    # Function to append a note in the first empty row
    def add_note(sheet_name, note_text):
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        sheet.cell(row=max_row + 2, column=1, value=note_text)  # Two rows after table

    # Add notes to "Tweets" and "Trump Tweets" sheets
    add_note("News Feeds",
             "*If it's empty, there are no related news to the keywords within 5 past days.")
    add_note("Tweets",
             "*Note that the tweets are not filtered on our side for relevancy, it only returns the most recent posts containing keywords provided by a user. So, knowing the nature of social networks, there might be irrelevant silly tweets.")
    add_note("Trump Tweets",
             "*Note that Trump tweets are not filtered by keywords, it is just the last 10 tweets of Trump.")

    # Save the modified workbook
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


def main():
    # st.set_page_config(page_title="MarketMuse – Your AI-powered muse for market inspiration",
    #                    #layout="wide"
    #                    )

    # Set up logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    # Title
    st.title("🧙‍♀️💡💰 MarketMuse")
    st.subheader("Your AI-powered muse for market inspiration")
    st.write("Scrape and analyze news & tweets based on your keywords.")

    # -------------------
    # User Input Section
    # -------------------
    st.header("🔍 Enter Search Parameters")
    keywords = st.text_area("Enter keywords (comma-separated):", placeholder="e.g. AI, technology, healthcare")
    business_type = st.text_input("Enter your business category:",
                                  placeholder="e.g. insurance, food service, healthcare")

    # -------------------
    # Data Fetching
    # -------------------
    if st.button("🚀 Fetch Data", use_container_width=True):
        # Clear previous session state values
        for key in ["response_json", "csv_data", "scrapes_excel", "all_suggested_niches", "chosen_niche"]:
            st.session_state.pop(key, None)

        keyword_list = [kw.strip() for kw in keywords.split(",") if kw.strip()]

        if keyword_list:
            with st.spinner("🔄 Fetching data... Please wait"):
                feed_results, tweets_data, trump_data = {}, {}, {}

                status = st.status("⏳ Processing queries...", expanded=True)

                # Fetch news feeds in parallel
                with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                    future_to_query = {executor.submit(fetch_feed, query): query for query in keyword_list}
                    for future in concurrent.futures.as_completed(future_to_query):
                        query = future_to_query[future]
                        try:
                            query_key, entries = future.result()
                            feed_results[query_key] = entries
                            status.update(label=f"✅ News fetched for {query}")
                        except Exception as exc:
                            logging.error(f"Query '{query}' generated an exception: {exc}")
                            status.update(label=f"⚠️ Error fetching news for {query}")

                # Scrape tweets from Nitter
                for query in keyword_list:
                    try:
                        tweets_data[query] = scrape_nitter(query)
                        status.update(label=f"✅ Tweets fetched for {query}")
                    except Exception as exc:
                        logging.error(f"Error scraping Nitter for '{query}': {exc}")
                status.update(label="⏳Scraping Trump's tweets")
                trump_data["Donald Trump Tweets"] = trump_scraper()
                status.update(label="✅ All data fetched! Passing it to AI for analysis...")

                combined_data = {
                    "news_feeds": feed_results,
                    "x_tweets": tweets_data,
                    "trump_data": trump_data
                }

                # Store processed data in session state
                st.session_state.response_json = choose_relevant_niches(combined_data, business_type)
                st.session_state.csv_data = convert_json_to_csv(st.session_state.response_json)
                st.session_state.scrapes_excel = save_scrapes_to_excel(combined_data)

                st.success("🎉 Data fetched and processed!")

        else:
            st.error("⚠️ Please enter at least one keyword.")

    # -------------------
    # Show Results
    # -------------------
    if "response_json" in st.session_state:
        st.header("📊 Data Insights")
        json_data = st.session_state.response_json

        if json_data:
            flat_data = []

            for category in json_data.get("Affected Business Categories", []):
                business_category = category.get("Business Category Name", "N/A")
                naic_code = category.get("NAIC Code", "N/A")
                potential_impact = category.get("Potential Impact", "N/A")

                suggested_niches = category.get("Suggested Niches", [])
                market_trends = category.get("Relevant Market Trends", [])

                flat_data.append({
                    "Business Category": business_category,
                    "NAIC Code": naic_code,
                    "Suggested Niches": suggested_niches, #"\n".join(suggested_niches) if suggested_niches else "N/A",
                    "Relevant Market Trends":  market_trends, #"\n".join(market_trends) if market_trends else "N/A",
                    "Potential Impact": potential_impact
                })

            if flat_data:
                df = pd.DataFrame(flat_data)
                # Apply CSS styling in Streamlit for wrapping text in dataframe
                st.markdown("""
                    <style>
                    div[data-testid="stDataFrame"] td {
                        white-space: pre-wrap;
                        word-wrap: break-word;
                    }
                    </style>
                """, unsafe_allow_html=True)

                # Display as an interactive table
                st.dataframe(df, use_container_width=True)
                # AgGrid(df, fit_columns_on_grid_load=True)
                # st.table(df)
                # st.data_editor(df, use_container_width=True, height=400)
            else:
                st.warning("No relevant business categories found.")
        # st.json(st.session_state.response_json)

        if "all_suggested_niches" not in st.session_state:
            st.session_state.all_suggested_niches = [
                niche
                for category in st.session_state.response_json["Affected Business Categories"]
                for niche in category.get("Suggested Niches", [])
            ]

        # -------------------
        # Enrichment Section
        # -------------------
        if st.session_state.all_suggested_niches:
            st.subheader("💡 Enrich Data with a Chosen Niche")
            with st.form("enrich_form"):
                chosen_niche = st.selectbox("Select a niche for further enrichment:",
                                            st.session_state.all_suggested_niches)
                submit_button = st.form_submit_button(label="🔍 Enrich Data")

            if submit_button:
                with st.spinner("🔄 Extracting relevant companies..."):
                    companies = run_graph(chosen_niche)
                    df = enrich(companies)
                    filtered_df = clean_company_list(df, chosen_niche)

                st.success("✅ Enrichment complete! See the relevant data below.")
                st.dataframe(filtered_df)

    # -------------------
    # Download Buttons
    # -------------------
    if "csv_data" in st.session_state and "scrapes_excel" in st.session_state:
        st.header("📥 Download Your Data")
        col1, col2 = st.columns(2)
        with col1:
            st.download_button("⬇️ Download CSV", st.session_state.csv_data, "scraped_data.csv", "text/csv",
                               use_container_width=True)
        with col2:
            st.download_button(
                "⬇️ Download Scraped News/Tweets",
                st.session_state.scrapes_excel,
                "scraped_data.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


if __name__ == "__main__":
    main()
